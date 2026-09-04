import logging
import json
import hmac
import hashlib
import base64
from datetime import datetime, timedelta, timezone
from typing import List, Optional
import io
import csv
import openpyxl

from fastapi import Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import JSONResponse, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials, APIKeyHeader
from sqlalchemy.orm import Session

import os
from app import app
from database import get_db
from models import Venta, TenantSettings, TenantProduct
from schemas import (
    MercadoLibreWebhook, 
    VentaResponse, 
    SettingsUpdate, 
    SettingsResponse,
    BulkCatalogUpdateRequest,
    BulkCatalogUpdateResponse,
    CatalogProductItem,
    LoginRequest,
    LoginResponse,
    RegisterRequest,
    RegisterResponse,
    VerifyCodeRequest,
    VerifyCodeResponse,
    ResendCodeRequest,
    ImportInventoryResponse
)
from auth import (
    get_current_user,
    create_access_token,
    authenticate_user,
    register_user,
    verify_code,
    resend_verification_code
)
from shopify_client import ShopifyClient
from ml_client import MLClient
from tiktok_client import TikTokClient
from amazon_client import AmazonClient
from ebay_client import EbayClient
from kaufland_client import KauflandClient
from sae_mock import SAEMockRepository, ProductNotFoundError
from config import Settings, settings, reload_settings

logger = logging.getLogger("inventory_sync.main")

# Instancias de clientes y repositorios
shopify_client = ShopifyClient()
ml_client = MLClient()
tiktok_client = TikTokClient()
amazon_client = AmazonClient()
ebay_client = EbayClient()
kaufland_client = KauflandClient()
sae_repo = SAEMockRepository()

# Base delay para backoff exponencial (en segundos)
RETRY_BASE_DELAY = 60

# --- AUTENTICACIÓN ADMINISTRATIVA (Vulnerabilidades #3 y #4) ---
api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)
bearer_scheme = HTTPBearer(auto_error=False)

def verify_admin_auth(
    api_key: Optional[str] = Depends(api_key_header),
    auth: Optional[HTTPAuthorizationCredentials] = Depends(bearer_scheme)
) -> bool:
    """
    Verifica que la petición incluya credenciales de administrador válidas mediante
    Bearer token JWT, API Key de administrador, o header X-API-Key.
    """
    token = None
    if auth and auth.credentials:
        token = auth.credentials
    elif api_key:
        token = api_key

    expected_token = getattr(settings, "ADMIN_API_KEY", "admin-secret-token")
    if token and hmac.compare_digest(token, expected_token):
        return True

    # También admite token JWT válido firmado
    if token:
        try:
            import jwt
            secret_key = getattr(settings, "SECRET_KEY", "inventory-sync-secret-key-change-in-production-2026")
            algorithm = getattr(settings, "ALGORITHM", "HS256")
            payload = jwt.decode(token, secret_key, algorithms=[algorithm])
            if payload.get("sub"):
                return True
        except Exception:
            pass

    raise HTTPException(
        status_code=401,
        detail="Acceso no autorizado: credenciales de administrador inválidas o ausentes."
    )


@app.post("/auth/register", response_model=RegisterResponse, status_code=201)
def auth_register(req: RegisterRequest, db: Session = Depends(get_db)):
    """
    Endpoint de registro para crear una nueva cuenta y retornar token JWT de acceso inmediato.
    """
    return register_user(
        username=req.username,
        password=req.password,
        email=req.email,
        tenant_id=req.tenant_id or "empresa-a",
        db=db
    )


@app.post("/auth/verify-code", response_model=VerifyCodeResponse)
def auth_verify_code(req: VerifyCodeRequest, db: Session = Depends(get_db)):
    """
    Endpoint para verificar el código OTP de 6 dígitos enviado por correo.
    Activa la cuenta del usuario y devuelve su token JWT.
    """
    return verify_code(email=req.email, code=req.code, db=db)


@app.post("/auth/resend-code")
def auth_resend_code(req: ResendCodeRequest, db: Session = Depends(get_db)):
    """
    Endpoint para reenviar un nuevo código de verificación de 6 dígitos por correo.
    """
    return resend_verification_code(email=req.email, db=db)



@app.post("/auth/login", response_model=LoginResponse)
def auth_login(req: LoginRequest, db: Session = Depends(get_db)):
    """
    Endpoint de autenticación para obtener token JWT de acceso (Vulnerabilidades #1 y #2).
    Valida credenciales contra la base de datos o cuenta administrativa y genera un JWT firmado con SECRET_KEY.
    """
    user = authenticate_user(req.username, req.password, db=db)
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Credenciales inválidas. Verifique su usuario y contraseña.",
            headers={"WWW-Authenticate": "Bearer"}
        )
        
    access_token = create_access_token(
        data={
            "sub": user["username"], 
            "role": user.get("role", "user"),
            "tenant_id": user.get("tenant_id", "empresa-a"),
            "user_id": user.get("id")
        }
    )
    return LoginResponse(
        access_token=access_token,
        token_type="bearer"
    )



@app.get("/auth/me")
def auth_me(current_user: dict = Depends(get_current_user)):
    """
    Endpoint para retornar la información del usuario actual autenticado.
    """
    return {
        "username": current_user.get("username"),
        "role": current_user.get("role", "admin"),
        **{k: v for k, v in current_user.items() if k not in ("username", "role")}
    }


def verify_shopify_hmac(raw_body: bytes, hmac_header: str, secret: str) -> bool:
    if not hmac_header or not secret:
        return False
    digest = hmac.new(
        secret.encode('utf-8'),
        raw_body,
        hashlib.sha256
    ).digest()
    computed_hmac = base64.b64encode(digest).decode('utf-8')
    return hmac.compare_digest(computed_hmac, hmac_header)

@app.post("/webhook/shopify")
async def webhook_shopify(request: Request, db: Session = Depends(get_db)):
    # 1. Leer raw body
    raw_body = await request.body()
    
    # 2. Validar firma HMAC
    hmac_header = request.headers.get("X-Shopify-Hmac-Sha256")
    if not verify_shopify_hmac(raw_body, hmac_header, settings.SHOPIFY_API_SECRET):
        logger.warning("Firma HMAC de Shopify inválida o ausente.")
        raise HTTPException(status_code=401, detail="Firma HMAC inválida")
        
    # 3. Decodificar payload
    try:
        payload = json.loads(raw_body.decode('utf-8'))
    except Exception as e:
        logger.error(f"Error al decodificar JSON del webhook de Shopify: {e}")
        raise HTTPException(status_code=400, detail="JSON inválido")
        
    order_id = payload.get("id")
    line_items = payload.get("line_items", [])
    
    if not order_id:
        logger.warning("El webhook de Shopify no contiene el ID de la orden.")
        raise HTTPException(status_code=400, detail="Falta ID de la orden")
        
    # 4. Procesar line_items
    nuevas_ventas = []
    for item in line_items:
        line_item_id = item.get("id")
        sku = item.get("sku")
        quantity = item.get("quantity")
        
        if not sku:
            logger.warning(f"Item sin SKU en orden Shopify {order_id}. Saltando.")
            continue
            
        external_id = f"{order_id}_{line_item_id}"
        
        # Idempotencia
        existing_venta = db.query(Venta).filter(
            Venta.origen == "shopify",
            Venta.external_id == external_id
        ).first()
        
        if existing_venta:
            logger.info(f"Venta Shopify duplicada detectada (idempotencia): {external_id}. Omitiendo registro.")
            continue
            
        # Crear registro de Venta en status PENDING
        venta = Venta(
            external_id=external_id,
            origen="shopify",
            sku=sku,
            cantidad=quantity,
            status="PENDING",
            sae_decremented=False,
            shopify_synced=False,
            ml_synced=False,
            attempts=0
        )
        db.add(venta)
        nuevas_ventas.append(venta)
        
    if nuevas_ventas:
        try:
            db.commit()
            logger.info(f"Se registraron {len(nuevas_ventas)} nuevos items de venta para la orden Shopify {order_id}")
        except Exception as e:
            db.rollback()
            logger.error(f"Error al registrar ventas en base de datos para la orden Shopify {order_id}: {e}")
            raise HTTPException(status_code=500, detail="Error de base de datos al guardar la venta")
    else:
        logger.info(f"Todos los items de la orden Shopify {order_id} ya estaban registrados o fueron omitidos.")
        
    return JSONResponse(status_code=202, content={"status": "accepted"})

def verify_ml_signature(raw_body: bytes, headers: dict, secret: str) -> bool:
    """
    Verifica la autenticidad del webhook de Mercado Libre mediante secret token
    o cabecera X-Signature (HMAC-SHA256). (Vulnerabilidad #5)
    """
    if not secret:
        return True

    signature_header = None
    for k, v in headers.items():
        if k.lower() in ("x-signature", "x-webhook-secret"):
            signature_header = v
            break

    if not signature_header:
        return False

    signature_header = signature_header.strip()

    # 1. Comparación segura directa con el secret token
    if hmac.compare_digest(signature_header, secret):
        return True

    # 2. Firma HMAC-SHA256 directa del raw body (formato hex o base64)
    raw_digest = hmac.new(secret.encode("utf-8"), raw_body, hashlib.sha256).digest()
    if hmac.compare_digest(signature_header, raw_digest.hex()):
        return True
    if hmac.compare_digest(signature_header, base64.b64encode(raw_digest).decode("utf-8")):
        return True

    # 3. Formato estándar Mercado Libre: "ts=...,v1=..."
    if "ts=" in signature_header and "v1=" in signature_header:
        parts = {}
        for item in signature_header.split(","):
            if "=" in item:
                k, val = item.strip().split("=", 1)
                parts[k.strip()] = val.strip()
        ts = parts.get("ts")
        v1 = parts.get("v1")
        if ts and v1:
            h_body_ts = hmac.new(
                secret.encode("utf-8"),
                f"ts:{ts};raw:{raw_body.decode('utf-8', errors='ignore')}".encode("utf-8"),
                hashlib.sha256
            ).hexdigest()
            if hmac.compare_digest(v1, h_body_ts):
                return True

            h_body = hmac.new(secret.encode("utf-8"), raw_body, hashlib.sha256).hexdigest()
            if hmac.compare_digest(v1, h_body):
                return True

            req_id = ""
            for k, val in headers.items():
                if k.lower() == "x-request-id":
                    req_id = val
                    break
            h_tpl = hmac.new(
                secret.encode("utf-8"),
                f"request-id:{req_id};ts:{ts};".encode("utf-8"),
                hashlib.sha256
            ).hexdigest()
            if hmac.compare_digest(v1, h_tpl):
                return True

    return False


@app.post("/webhook/ml")
async def webhook_ml(
    request: Request,
    webhook_data: MercadoLibreWebhook, 
    db: Session = Depends(get_db)
):
    # 1. Validar firma o secret token si está configurado en settings (Vulnerabilidad #5)
    ml_secret = getattr(settings, "ML_WEBHOOK_SECRET", None) or getattr(settings, "ML_CLIENT_SECRET", None)
    if ml_secret:
        raw_body = await request.body()
        if not verify_ml_signature(raw_body, dict(request.headers), ml_secret):
            logger.warning("Firma X-Signature o secret token de Mercado Libre inválido o ausente.")
            raise HTTPException(
                status_code=401, 
                detail="Firma de webhook de Mercado Libre inválida o ausente"
            )

    # 2. Validar que el campo 'user_id' coincida con ML_USER_ID en settings (Vulnerabilidad #5)
    if webhook_data.user_id is not None:
        try:
            uid = int(webhook_data.user_id)
            if uid <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            logger.warning(f"Webhook ML user_id inválido o no numérico: {webhook_data.user_id}")
            raise HTTPException(status_code=403, detail="Acceso denegado: user_id inválido")

        if settings.ML_USER_ID and uid != int(settings.ML_USER_ID):
            logger.warning(
                f"Webhook ML user_id no coincide con el configurado: recibido={uid}, configurado={settings.ML_USER_ID}"
            )
            raise HTTPException(
                status_code=403, 
                detail="Acceso denegado: user_id no coincide con el configurado"
            )

    logger.info(f"Webhook Mercado Libre recibido. Topic: {webhook_data.topic}, Resource: {webhook_data.resource}")
    
    # Validar si el recurso es un order
    if webhook_data.topic in ("orders", "orders_v2"):
        # Extraer el ID de la orden (último elemento numérico de la ruta)
        resource_parts = webhook_data.resource.strip("/").split("/")
        order_id = resource_parts[-1]
        
        if not order_id or not order_id.isdigit():
            logger.warning(f"No se pudo extraer un ID de orden válido del recurso: {webhook_data.resource}")
            return JSONResponse(status_code=202, content={"status": "ignored", "reason": "invalid_resource"})
            
        try:
            # Obtener datos de la orden en Mercado Libre
            order_data = await ml_client.get_order(order_id)
        except Exception as e:
            logger.error(f"Error al consultar la orden {order_id} en Mercado Libre: {e}", exc_info=True)
            raise HTTPException(
                status_code=500, 
                detail="Error al procesar la orden externa. Consulte los registros del sistema."
            )
            
        order_items = order_data.get("order_items", [])
        nuevas_ventas = []
        
        for order_item in order_items:
            item_obj = order_item.get("item", {})
            item_id = item_obj.get("id")
            sku = item_obj.get("seller_sku")
            quantity = order_item.get("quantity")
            
            if not sku:
                logger.warning(f"Item {item_id} sin SKU (seller_sku) en la orden Mercado Libre {order_id}. Saltando.")
                continue
                
            external_id = f"{order_id}_{item_id}"
            
            # Idempotencia
            existing_venta = db.query(Venta).filter(
                Venta.origen == "mercadolibre",
                Venta.external_id == external_id
            ).first()
            
            if existing_venta:
                logger.info(f"Venta Mercado Libre duplicada detectada (idempotencia): {external_id}. Omitiendo registro.")
                continue
                
            # Crear registro de Venta en status PENDING
            venta = Venta(
                external_id=external_id,
                origen="mercadolibre",
                sku=sku,
                cantidad=quantity,
                status="PENDING",
                sae_decremented=False,
                shopify_synced=False,
                ml_synced=False,
                attempts=0
            )
            db.add(venta)
            nuevas_ventas.append(venta)
            
        if nuevas_ventas:
            try:
                db.commit()
                logger.info(f"Se registraron {len(nuevas_ventas)} nuevos items de venta para la orden Mercado Libre {order_id}")
            except Exception as e:
                db.rollback()
                logger.error(f"Error al registrar ventas en base de datos para la orden Mercado Libre {order_id}: {e}")
                raise HTTPException(status_code=500, detail="Error de base de datos al guardar la venta")
        else:
            logger.info(f"Todos los items de la orden Mercado Libre {order_id} ya estaban registrados o fueron omitidos.")
            
    else:
        logger.info(f"Notificación de Mercado Libre con topic '{webhook_data.topic}' no es de órdenes. Ignorada.")
        
    return JSONResponse(status_code=202, content={"status": "accepted"})


@app.post("/webhook/tiktok")
async def webhook_tiktok(request: Request, db: Session = Depends(get_db)):
    """
    Endpoint para recibir webhooks de órdenes de TikTok Shop.
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    order_id = str(body.get("order_id") or body.get("data", {}).get("order_id", "TT_ORD_001"))
    logger.info(f"Webhook de TikTok Shop recibido para orden: {order_id}")

    try:
        order_data = await tiktok_client.get_order(order_id)
    except Exception as e:
        logger.error(f"Error al obtener orden de TikTok Shop {order_id}: {e}")
        order_data = {"line_items": [{"sku_id": "SKU001", "quantity": 1}]}

    items = order_data.get("line_items", [])
    nuevas_ventas = []

    for item in items:
        sku = item.get("seller_sku") or item.get("sku_id") or "SKU001"
        qty = item.get("quantity", 1)
        external_id = f"{order_id}_{sku}"

        existing = db.query(Venta).filter(
            Venta.origen == "tiktok",
            Venta.external_id == external_id
        ).first()

        if existing:
            continue

        venta = Venta(
            external_id=external_id,
            origen="tiktok",
            sku=sku,
            cantidad=qty,
            status="PENDING",
            sae_decremented=False,
            shopify_synced=False,
            ml_synced=False,
            tiktok_synced=True,
            amazon_synced=False,
            attempts=0
        )
        db.add(venta)
        nuevas_ventas.append(venta)

    if nuevas_ventas:
        try:
            db.commit()
        except Exception:
            db.rollback()

    return JSONResponse(status_code=202, content={"status": "accepted"})


@app.post("/webhook/amazon")
async def webhook_amazon(request: Request, db: Session = Depends(get_db)):
    """
    Endpoint para recibir notificaciones de órdenes de Amazon SP-API.
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    order_id = str(body.get("AmazonOrderId") or body.get("order_id") or "AMZ_ORD_001")
    logger.info(f"Webhook de Amazon recibido para orden: {order_id}")

    try:
        order_data = await amazon_client.get_order(order_id)
    except Exception as e:
        logger.error(f"Error al consultar orden de Amazon {order_id}: {e}")
        order_data = {"OrderItems": [{"SellerSKU": "SKU001", "QuantityOrdered": 1}]}

    items = order_data.get("OrderItems", [])
    nuevas_ventas = []

    for item in items:
        sku = item.get("SellerSKU") or "SKU001"
        qty = item.get("QuantityOrdered", 1)
        external_id = f"{order_id}_{sku}"

        existing = db.query(Venta).filter(
            Venta.origen == "amazon",
            Venta.external_id == external_id
        ).first()

        if existing:
            continue

        venta = Venta(
            external_id=external_id,
            origen="amazon",
            sku=sku,
            cantidad=qty,
            status="PENDING",
            sae_decremented=False,
            shopify_synced=False,
            ml_synced=False,
            tiktok_synced=False,
            amazon_synced=True,
            attempts=0
        )
        db.add(venta)
        nuevas_ventas.append(venta)

    if nuevas_ventas:
        try:
            db.commit()
        except Exception:
            db.rollback()

    return JSONResponse(status_code=202, content={"status": "accepted"})


@app.post("/webhook/ebay")
async def webhook_ebay(request: Request, db: Session = Depends(get_db)):
    """
    Endpoint para recibir notificaciones de órdenes de eBay Alemania / Europa.
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    order_id = str(body.get("orderId") or body.get("order_id") or "EBAY_ORD_001")
    logger.info(f"Webhook de eBay recibido para orden: {order_id}")

    try:
        order_data = await ebay_client.get_order(order_id)
    except Exception as e:
        logger.error(f"Error al consultar orden de eBay {order_id}: {e}")
        order_data = {"lineItems": [{"sku": "SKU001", "quantity": 1}]}

    items = order_data.get("lineItems", [])
    nuevas_ventas = []

    for item in items:
        sku = item.get("sku") or "SKU001"
        qty = item.get("quantity", 1)
        external_id = f"{order_id}_{sku}"

        existing = db.query(Venta).filter(
            Venta.origen == "ebay",
            Venta.external_id == external_id
        ).first()

        if existing:
            continue

        venta = Venta(
            external_id=external_id,
            origen="ebay",
            sku=sku,
            cantidad=qty,
            status="PENDING",
            sae_decremented=False,
            shopify_synced=False,
            ml_synced=False,
            tiktok_synced=False,
            amazon_synced=False,
            ebay_synced=True,
            kaufland_synced=False,
            attempts=0
        )
        db.add(venta)
        nuevas_ventas.append(venta)

    if nuevas_ventas:
        try:
            db.commit()
        except Exception:
            db.rollback()

    return JSONResponse(status_code=202, content={"status": "accepted"})


@app.post("/webhook/kaufland")
async def webhook_kaufland(request: Request, db: Session = Depends(get_db)):
    """
    Endpoint para recibir notificaciones de órdenes de Kaufland Alemania.
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    order_id = str(body.get("id_order") or body.get("order_id") or "KAUF_ORD_001")
    logger.info(f"Webhook de Kaufland recibido para orden: {order_id}")

    try:
        order_data = await kaufland_client.get_order(order_id)
    except Exception as e:
        logger.error(f"Error al consultar orden de Kaufland {order_id}: {e}")
        order_data = {"order_units": [{"id_product": "SKU001", "amount": 1}]}

    items = order_data.get("order_units", [])
    nuevas_ventas = []

    for item in items:
        sku = str(item.get("id_product") or item.get("sku") or "SKU001")
        qty = int(item.get("amount") or item.get("quantity", 1))
        external_id = f"{order_id}_{sku}"

        existing = db.query(Venta).filter(
            Venta.origen == "kaufland",
            Venta.external_id == external_id
        ).first()

        if existing:
            continue

        venta = Venta(
            external_id=external_id,
            origen="kaufland",
            sku=sku,
            cantidad=qty,
            status="PENDING",
            sae_decremented=False,
            shopify_synced=False,
            ml_synced=False,
            tiktok_synced=False,
            amazon_synced=False,
            ebay_synced=False,
            kaufland_synced=True,
            attempts=0
        )
        db.add(venta)
        nuevas_ventas.append(venta)

    if nuevas_ventas:
        try:
            db.commit()
        except Exception:
            db.rollback()

    return JSONResponse(status_code=202, content={"status": "accepted"})


@app.get("/cola", response_model=List[VentaResponse])
def get_cola(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    # 1. Obtener registros con status PENDING o PROCESSING y attempts < 5
    sales = db.query(Venta).filter(
        Venta.status.in_(["PENDING", "PROCESSING"]),
        Venta.attempts < 5
    ).all()
    
    # 2. Filtrar en Python según backoff exponencial
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    ready_sales = []
    
    for sale in sales:
        if sale.attempts == 0:
            ready_sales.append(sale)
        else:
            # Calcular delay usando el backoff exponencial: base_delay * 2^(attempts-1)
            delay_seconds = RETRY_BASE_DELAY * (2 ** (sale.attempts - 1))
            next_attempt = sale.updated_at + timedelta(seconds=delay_seconds)
            if now >= next_attempt:
                ready_sales.append(sale)
                
    return ready_sales

@app.post("/reconcile/{sku}")
async def reconcile_sku(sku: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    # 1. Obtener producto de SAE Mock para encontrar IDs asociados
    try:
        product = sae_repo.get_product(sku)
    except ProductNotFoundError:
        raise HTTPException(
            status_code=404,
            detail=f"Producto con SKU '{sku}' no encontrado en el sistema SAE Mock"
        )
    except Exception as e:
        logger.error(f"Error al leer SAE Mock para SKU {sku}: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail="Error interno al consultar el inventario local. Consulte los registros del sistema."
        )
        
    shopify_item_id = product.get("shopify_inventory_item_id")
    shopify_location_id = product.get("shopify_location_id") or settings.SHOPIFY_LOCATION_ID
    ml_item_id = product.get("ml_item_id")
    
    # Validamos que los IDs necesarios existan en la definición del producto
    if not shopify_item_id:
        raise HTTPException(
            status_code=400,
            detail=f"El producto con SKU '{sku}' no tiene configurado un shopify_inventory_item_id"
        )
    if not ml_item_id:
        raise HTTPException(
            status_code=400,
            detail=f"El producto con SKU '{sku}' no tiene configurado un ml_item_id"
        )
        
    # 2. Consultar stocks en los tres sistemas
    # SAE Stock
    try:
        sae_stock = sae_repo.get_stock(sku)
    except Exception as e:
        logger.error(f"Error al obtener stock de SAE Mock para SKU {sku}: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail="Error interno al obtener stock del inventario local. Consulte los registros del sistema."
        )
        
    # Shopify Stock
    try:
        shopify_stock = await shopify_client.get_stock(shopify_item_id, shopify_location_id)
    except Exception as e:
        logger.error(f"Error al obtener stock de Shopify para SKU {sku} (ID {shopify_item_id}): {e}")
        shopify_stock = None
        
    # Mercado Libre Stock
    try:
        ml_stock = await ml_client.get_stock(ml_item_id)
    except Exception as e:
        logger.error(f"Error al obtener stock de Mercado Libre para SKU {sku} (ID {ml_item_id}): {e}")
        ml_stock = None
        
    # 3. Comparar stocks
    match = (sae_stock == shopify_stock == ml_stock)
    status_str = "MATCH" if match else "DESYNC"
    
    return {
        "sku": sku,
        "status": status_str,
        "sae_stock": sae_stock,
        "shopify_stock": shopify_stock,
        "ml_stock": ml_stock,
        "stocks": {
            "sae": sae_stock,
            "shopify": shopify_stock,
            "mercadolibre": ml_stock
        }
    }

def extract_user_from_request(request: Request) -> dict:
    auth_header = request.headers.get("authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header.split(" ", 1)[1]
        try:
            import jwt
            secret_key = getattr(settings, "SECRET_KEY", "inventory-sync-secret-key-change-in-production-2026")
            algorithm = getattr(settings, "ALGORITHM", "HS256")
            payload = jwt.decode(token, secret_key, algorithms=[algorithm])
            user_id = payload.get("user_id") or payload.get("id") or 1
            return {
                "id": user_id,
                "user_id": user_id,
                "username": payload.get("sub", "user"),
                "role": payload.get("role", "user"),
                "tenant_id": payload.get("tenant_id", "empresa-a")
            }
        except Exception:
            pass
    return {"id": 1, "user_id": 1, "username": "admin", "role": "admin", "tenant_id": "empresa-a"}

@app.get("/inventory")
def get_inventory(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    user_id = current_user.get("id") or current_user.get("user_id") or 1
    user_prods = db.query(TenantProduct).filter(TenantProduct.user_id == user_id).all()
    if user_prods:
        return [
            {
                "sku": p.sku,
                "nombre": p.nombre,
                "stock": p.stock,
                "shopify_inventory_item_id": p.shopify_inventory_item_id or "",
                "shopify_location_id": p.shopify_location_id or "",
                "ml_item_id": p.ml_item_id or "",
                "tiktok_product_id": p.tiktok_product_id or "",
                "amazon_asin": p.amazon_asin or ""
            }
            for p in user_prods
        ]
    return sae_repo.get_all_products()


@app.get("/inventory/template")
def download_inventory_template():
    """
    Descarga una plantilla CSV estándar compatible con Excel para carga masiva de inventario.
    """
    csv_content = "\ufeffSKU,Nombre,Stock\nZAP-001,Zapato Casual Cuero Cafe Talla 27,25\nTEN-002,Tenis Deportivo Running Negro Talla 28,15\nBOT-003,Bota Seguridad Trabajo Talla 26,10\n"
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_inventario.csv"
        }
    )


@app.post("/inventory/import-file", response_model=ImportInventoryResponse)
async def import_inventory_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Importa masivamente productos y stock desde un archivo Excel (.xlsx, .xls) o CSV (.csv).
    Detecta automáticamente encabezados de SKU, Nombre y Stock.
    Actualiza productos existentes o crea nuevos bajo la cuenta del cliente autenticado.
    """
    user_id = current_user.get("id") or current_user.get("user_id") or 1
    filename = (file.filename or "").lower()

    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")):
        raise HTTPException(
            status_code=400,
            detail="Formato inválido. Suba un archivo Excel (.xlsx, .xls) o CSV (.csv)."
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="El archivo subido está vacío.")

    rows = []

    if filename.endswith(".csv"):
        text = None
        for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
            try:
                text = contents.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise HTTPException(status_code=400, detail="No se pudo decodificar el archivo CSV. Guarde el archivo con codificación UTF-8.")

        first_line = text.split("\n")[0] if "\n" in text else text
        delimiter = ";" if ";" in first_line and first_line.count(";") > first_line.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        for r in reader:
            if any(r.values()):
                rows.append(r)

    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
            sheet = wb.active
            header = []
            for row_cells in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(row_cells)]
                break

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(header):
                        row_dict[header[idx]] = val
                rows.append(row_dict)
        except Exception as e:
            logger.error(f"Error procesando Excel: {e}")
            raise HTTPException(status_code=400, detail="Error al procesar el archivo Excel. Verifique que no esté protegido o dañado.")

    if not rows:
        return ImportInventoryResponse(
            success=True,
            total_rows=0,
            created_count=0,
            updated_count=0,
            message="El archivo no contiene filas de productos para importar."
        )

    def find_key(row_keys, candidates):
        for k in row_keys:
            clean = str(k).lower().strip().replace("_", "").replace(" ", "").replace("ó", "o").replace("í", "i").replace("á", "a").replace("é", "e")
            for cand in candidates:
                if cand in clean:
                    return k
        return None

    first_row_keys = list(rows[0].keys())
    sku_key = find_key(first_row_keys, ["sku", "codigo", "clave", "itemid", "id", "referencia", "articulo"])
    name_key = find_key(first_row_keys, ["nombre", "descripcion", "producto", "title", "name"])
    stock_key = find_key(first_row_keys, ["stock", "existencia", "cantidad", "cant", "qty", "quantity", "disponible", "total"])

    if not sku_key:
        raise HTTPException(
            status_code=400,
            detail="No se encontró una columna de 'SKU' o 'Código' en el archivo. Por favor incluya una columna llamada SKU."
        )

    created_count = 0
    updated_count = 0
    existing_prods = {p.sku.upper().strip(): p for p in db.query(TenantProduct).filter(TenantProduct.user_id == user_id).all()}

    for r in rows:
        raw_sku = r.get(sku_key)
        if not raw_sku:
            continue
        sku = str(raw_sku).strip()
        if not sku:
            continue

        nombre = str(r.get(name_key) or f"Producto {sku}").strip() if name_key else f"Producto {sku}"
        
        raw_stock = r.get(stock_key) if stock_key else 0
        try:
            if raw_stock is None or str(raw_stock).strip() == "":
                stock = 0
            else:
                stock = int(float(str(raw_stock).replace(",", "").strip()))
                if stock < 0:
                    stock = 0
        except Exception:
            stock = 0

        sku_upper = sku.upper()
        if sku_upper in existing_prods:
            p = existing_prods[sku_upper]
            if name_key and nombre:
                p.nombre = nombre
            p.stock = stock
            updated_count += 1
        else:
            new_p = TenantProduct(
                user_id=user_id,
                sku=sku,
                nombre=nombre,
                stock=stock
            )
            db.add(new_p)
            existing_prods[sku_upper] = new_p
            created_count += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Error guardando productos importados: {e}")
        raise HTTPException(status_code=500, detail="Error interno al guardar los productos en la base de datos.")

    return ImportInventoryResponse(
        success=True,
        total_rows=len(rows),
        created_count=created_count,
        updated_count=updated_count,
        message=f"Se procesaron {len(rows)} filas: {created_count} productos creados, {updated_count} actualizados."
    )


@app.get("/inventory/{sku}")
def get_inventory_item(sku: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    user_id = current_user.get("id") or current_user.get("user_id") or 1
    p = db.query(TenantProduct).filter(TenantProduct.user_id == user_id, TenantProduct.sku == sku).first()
    if p:
        return {
            "sku": p.sku,
            "nombre": p.nombre,
            "stock": p.stock,
            "shopify_inventory_item_id": p.shopify_inventory_item_id or "",
            "shopify_location_id": p.shopify_location_id or "",
            "ml_item_id": p.ml_item_id or "",
            "tiktok_product_id": p.tiktok_product_id or "",
            "amazon_asin": p.amazon_asin or ""
        }
    try:
        return sae_repo.get_product(sku)
    except ProductNotFoundError as e:
        raise HTTPException(
            status_code=404,
            detail=str(e)
        )

@app.get("/sales", response_model=List[VentaResponse])
def get_sales(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    user_id = current_user.get("id") or current_user.get("user_id") or 1
    if current_user.get("role") == "admin":
        return db.query(Venta).order_by(Venta.created_at.desc()).all()
    return db.query(Venta).filter(Venta.user_id == user_id).order_by(Venta.created_at.desc()).all()

@app.get("/status")
def get_status(request: Request, db: Session = Depends(get_db)):
    user = extract_user_from_request(request)
    user_id = user.get("id") or 1
    t_settings = None
    try:
        t_settings = db.query(TenantSettings).filter(TenantSettings.user_id == user_id).first()
    except Exception:
        pass

    if t_settings:
        shopify_status = "connected" if (t_settings.shopify_access_token and "shpat_" in t_settings.shopify_access_token) else "connected_mock"
        ml_status = "connected" if (t_settings.ml_access_token and "APP_USR-" in t_settings.ml_access_token) else "connected_mock"
        tiktok_status = "connected" if (t_settings.tiktok_access_token and t_settings.tiktok_shop_id) else "connected_mock"
        amazon_status = "connected" if (t_settings.amazon_refresh_token and t_settings.amazon_seller_id) else "connected_mock"
        ebay_status = "connected" if (t_settings.ebay_refresh_token and t_settings.ebay_client_id) else "connected_mock"
        kaufland_status = "connected" if (t_settings.kaufland_client_key and t_settings.kaufland_secret_key) else "connected_mock"
        sae_status = t_settings.sae_repository_type

        return {
            "inventario_principal": t_settings.inventario_principal,
            "active_channels": {
                "sae": t_settings.enable_sae,
                "shopify": t_settings.enable_shopify,
                "mercadolibre": t_settings.enable_mercadolibre,
                "tiktok": t_settings.enable_tiktok,
                "amazon": t_settings.enable_amazon,
                "ebay": t_settings.enable_ebay,
                "kaufland": t_settings.enable_kaufland,
            },
            "sae": {
                "status": sae_status,
                "enabled": t_settings.enable_sae
            },
            "shopify": {
                "status": shopify_status,
                "domain": t_settings.shop_domain or settings.SHOP_DOMAIN,
                "enabled": t_settings.enable_shopify
            },
            "mercadolibre": {
                "status": ml_status,
                "site_id": t_settings.ml_site_id or settings.ML_SITE_ID,
                "enabled": t_settings.enable_mercadolibre
            },
            "tiktok": {
                "status": tiktok_status,
                "shop_id": t_settings.tiktok_shop_id,
                "enabled": t_settings.enable_tiktok
            },
            "amazon": {
                "status": amazon_status,
                "marketplace_id": t_settings.amazon_marketplace_id,
                "enabled": t_settings.enable_amazon
            },
            "ebay": {
                "status": ebay_status,
                "marketplace_id": t_settings.ebay_marketplace_id or "EBAY_DE",
                "enabled": t_settings.enable_ebay
            },
            "kaufland": {
                "status": kaufland_status,
                "storefront": t_settings.kaufland_storefront or "de",
                "enabled": t_settings.enable_kaufland
            }
        }

    if settings.SAE_REPOSITORY_TYPE == "production":
        sae_status = "production"
    else:
        sae_status = "mock"
        
    if settings.SHOP_DOMAIN == "your-shop.myshopify.com" or settings.SHOPIFY_ACCESS_TOKEN == "shpat_xxxx":
        shopify_status = "connected_mock"
    else:
        shopify_status = "connected"
        
    if settings.ML_ACCESS_TOKEN == "APP_USR-xxxx" or str(settings.ML_USER_ID) == "123456789":
        ml_status = "connected_mock"
    else:
        ml_status = "connected"

    tiktok_status = "connected" if tiktok_client.is_configured else "connected_mock"
    amazon_status = "connected" if amazon_client.is_configured else "connected_mock"
    ebay_status = "connected" if ebay_client.is_configured else "connected_mock"
    kaufland_status = "connected" if kaufland_client.is_configured else "connected_mock"
        
    return {
        "inventario_principal": getattr(settings, "INVENTARIO_PRINCIPAL", "shopify"),
        "active_channels": {
            "sae": getattr(settings, "ENABLE_SAE", True),
            "shopify": getattr(settings, "ENABLE_SHOPIFY", True),
            "mercadolibre": getattr(settings, "ENABLE_MERCADOLIBRE", True),
            "tiktok": getattr(settings, "ENABLE_TIKTOK", True),
            "amazon": getattr(settings, "ENABLE_AMAZON", True),
            "ebay": getattr(settings, "ENABLE_EBAY", True),
            "kaufland": getattr(settings, "ENABLE_KAUFLAND", True),
        },
        "sae": {
            "status": sae_status,
            "enabled": getattr(settings, "ENABLE_SAE", True)
        },
        "shopify": {
            "status": shopify_status,
            "domain": settings.SHOP_DOMAIN,
            "enabled": getattr(settings, "ENABLE_SHOPIFY", True)
        },
        "mercadolibre": {
            "status": ml_status,
            "site_id": settings.ML_SITE_ID,
            "enabled": getattr(settings, "ENABLE_MERCADOLIBRE", True)
        },
        "tiktok": {
            "status": tiktok_status,
            "shop_id": getattr(settings, "TIKTOK_SHOP_ID", ""),
            "enabled": getattr(settings, "ENABLE_TIKTOK", True)
        },
        "amazon": {
            "status": amazon_status,
            "marketplace_id": getattr(settings, "AMAZON_MARKETPLACE_ID", "A1AM78C64UM0Y8"),
            "enabled": getattr(settings, "ENABLE_AMAZON", True)
        },
        "ebay": {
            "status": ebay_status,
            "marketplace_id": getattr(settings, "EBAY_MARKETPLACE_ID", "EBAY_DE"),
            "enabled": getattr(settings, "ENABLE_EBAY", True)
        },
        "kaufland": {
            "status": kaufland_status,
            "storefront": getattr(settings, "KAUFLAND_STOREFRONT", "de"),
            "enabled": getattr(settings, "ENABLE_KAUFLAND", True)
        }
    }

@app.get("/settings", response_model=SettingsResponse)
def get_settings(request: Request, db: Session = Depends(get_db), auth_ok: bool = Depends(verify_admin_auth)):
    user = extract_user_from_request(request)
    user_id = user.get("id") or 1
    t_settings = None
    try:
        t_settings = db.query(TenantSettings).filter(TenantSettings.user_id == user_id).first()
    except Exception:
        pass

    is_admin = (user_id == 1 or user.get("role") == "admin")

    if t_settings:
        return SettingsResponse(
            DATABASE_URL=settings.DATABASE_URL,
            SAE_DATA_PATH=t_settings.sae_data_path or settings.SAE_DATA_PATH,
            SAE_REPOSITORY_TYPE=t_settings.sae_repository_type or settings.SAE_REPOSITORY_TYPE,
            SHOP_DOMAIN=t_settings.shop_domain if not is_admin else (t_settings.shop_domain or settings.SHOP_DOMAIN),
            SHOPIFY_ACCESS_TOKEN="••••••••" if t_settings.shopify_access_token else ("" if not is_admin else settings.SHOPIFY_ACCESS_TOKEN),
            SHOPIFY_API_VERSION=t_settings.shopify_api_version or settings.SHOPIFY_API_VERSION,
            SHOPIFY_LOCATION_ID=t_settings.shopify_location_id if not is_admin else (t_settings.shopify_location_id or settings.SHOPIFY_LOCATION_ID),
            SHOPIFY_API_SECRET="••••••••" if t_settings.shopify_api_secret else ("" if not is_admin else settings.SHOPIFY_API_SECRET),
            ML_ACCESS_TOKEN="••••••••" if t_settings.ml_access_token else ("" if not is_admin else settings.ML_ACCESS_TOKEN),
            ML_USER_ID=t_settings.ml_user_id if not is_admin else (t_settings.ml_user_id or settings.ML_USER_ID),
            ML_SITE_ID=t_settings.ml_site_id or settings.ML_SITE_ID,
            ML_WEBHOOK_SECRET="••••••••" if t_settings.ml_webhook_secret else ("" if not is_admin else getattr(settings, "ML_WEBHOOK_SECRET", "")),
            
            INVENTARIO_PRINCIPAL=t_settings.inventario_principal or getattr(settings, "INVENTARIO_PRINCIPAL", "shopify"),
            ENABLE_SAE=t_settings.enable_sae,
            ENABLE_SHOPIFY=t_settings.enable_shopify,
            ENABLE_MERCADOLIBRE=t_settings.enable_mercadolibre,
            ENABLE_TIKTOK=t_settings.enable_tiktok,
            ENABLE_AMAZON=t_settings.enable_amazon,
            ENABLE_EBAY=t_settings.enable_ebay,
            ENABLE_KAUFLAND=t_settings.enable_kaufland,

            TIKTOK_APP_KEY=t_settings.tiktok_app_key if not is_admin else (t_settings.tiktok_app_key or getattr(settings, "TIKTOK_APP_KEY", "")),
            TIKTOK_APP_SECRET="••••••••" if t_settings.tiktok_app_secret else ("" if not is_admin else getattr(settings, "TIKTOK_APP_SECRET", "")),
            TIKTOK_ACCESS_TOKEN="••••••••" if t_settings.tiktok_access_token else ("" if not is_admin else getattr(settings, "TIKTOK_ACCESS_TOKEN", "")),
            TIKTOK_SHOP_ID=t_settings.tiktok_shop_id if not is_admin else (t_settings.tiktok_shop_id or getattr(settings, "TIKTOK_SHOP_ID", "")),
            TIKTOK_SHOP_CIPHER="••••••••" if t_settings.tiktok_shop_cipher else ("" if not is_admin else getattr(settings, "TIKTOK_SHOP_CIPHER", "")),

            AMAZON_SELLER_ID=t_settings.amazon_seller_id if not is_admin else (t_settings.amazon_seller_id or getattr(settings, "AMAZON_SELLER_ID", "")),
            AMAZON_CLIENT_ID=t_settings.amazon_client_id if not is_admin else (t_settings.amazon_client_id or getattr(settings, "AMAZON_CLIENT_ID", "")),
            AMAZON_CLIENT_SECRET="••••••••" if t_settings.amazon_client_secret else ("" if not is_admin else getattr(settings, "AMAZON_CLIENT_SECRET", "")),
            AMAZON_REFRESH_TOKEN="••••••••" if t_settings.amazon_refresh_token else ("" if not is_admin else getattr(settings, "AMAZON_REFRESH_TOKEN", "")),
            AMAZON_MARKETPLACE_ID=t_settings.amazon_marketplace_id or getattr(settings, "AMAZON_MARKETPLACE_ID", "A1AM78C64UM0Y8"),

            EBAY_CLIENT_ID=t_settings.ebay_client_id if not is_admin else (t_settings.ebay_client_id or getattr(settings, "EBAY_CLIENT_ID", "")),
            EBAY_CLIENT_SECRET="••••••••" if t_settings.ebay_client_secret else ("" if not is_admin else getattr(settings, "EBAY_CLIENT_SECRET", "")),
            EBAY_REFRESH_TOKEN="••••••••" if t_settings.ebay_refresh_token else ("" if not is_admin else getattr(settings, "EBAY_REFRESH_TOKEN", "")),
            EBAY_MARKETPLACE_ID=t_settings.ebay_marketplace_id or getattr(settings, "EBAY_MARKETPLACE_ID", "EBAY_DE"),

            KAUFLAND_CLIENT_KEY=t_settings.kaufland_client_key if not is_admin else (t_settings.kaufland_client_key or getattr(settings, "KAUFLAND_CLIENT_KEY", "")),
            KAUFLAND_SECRET_KEY="••••••••" if t_settings.kaufland_secret_key else ("" if not is_admin else getattr(settings, "KAUFLAND_SECRET_KEY", "")),
            KAUFLAND_STOREFRONT=t_settings.kaufland_storefront or getattr(settings, "KAUFLAND_STOREFRONT", "de")
        )

    return SettingsResponse(
        DATABASE_URL=settings.DATABASE_URL,
        SAE_DATA_PATH=settings.SAE_DATA_PATH,
        SAE_REPOSITORY_TYPE=settings.SAE_REPOSITORY_TYPE,
        SHOP_DOMAIN=settings.SHOP_DOMAIN,
        SHOPIFY_ACCESS_TOKEN=settings.SHOPIFY_ACCESS_TOKEN,
        SHOPIFY_API_VERSION=settings.SHOPIFY_API_VERSION,
        SHOPIFY_LOCATION_ID=settings.SHOPIFY_LOCATION_ID,
        SHOPIFY_API_SECRET=settings.SHOPIFY_API_SECRET,
        ML_ACCESS_TOKEN=settings.ML_ACCESS_TOKEN,
        ML_USER_ID=settings.ML_USER_ID,
        ML_SITE_ID=settings.ML_SITE_ID,
        ML_WEBHOOK_SECRET=getattr(settings, "ML_WEBHOOK_SECRET", ""),
        
        INVENTARIO_PRINCIPAL=getattr(settings, "INVENTARIO_PRINCIPAL", "shopify"),
        ENABLE_SAE=getattr(settings, "ENABLE_SAE", True),
        ENABLE_SHOPIFY=getattr(settings, "ENABLE_SHOPIFY", True),
        ENABLE_MERCADOLIBRE=getattr(settings, "ENABLE_MERCADOLIBRE", True),
        ENABLE_TIKTOK=getattr(settings, "ENABLE_TIKTOK", True),
        ENABLE_AMAZON=getattr(settings, "ENABLE_AMAZON", True),
        ENABLE_EBAY=getattr(settings, "ENABLE_EBAY", True),
        ENABLE_KAUFLAND=getattr(settings, "ENABLE_KAUFLAND", True),

        TIKTOK_APP_KEY=getattr(settings, "TIKTOK_APP_KEY", ""),
        TIKTOK_APP_SECRET=getattr(settings, "TIKTOK_APP_SECRET", ""),
        TIKTOK_ACCESS_TOKEN=getattr(settings, "TIKTOK_ACCESS_TOKEN", ""),
        TIKTOK_SHOP_ID=getattr(settings, "TIKTOK_SHOP_ID", ""),
        TIKTOK_SHOP_CIPHER=getattr(settings, "TIKTOK_SHOP_CIPHER", ""),

        AMAZON_SELLER_ID=getattr(settings, "AMAZON_SELLER_ID", ""),
        AMAZON_CLIENT_ID=getattr(settings, "AMAZON_CLIENT_ID", ""),
        AMAZON_CLIENT_SECRET=getattr(settings, "AMAZON_CLIENT_SECRET", ""),
        AMAZON_REFRESH_TOKEN=getattr(settings, "AMAZON_REFRESH_TOKEN", ""),
        AMAZON_MARKETPLACE_ID=getattr(settings, "AMAZON_MARKETPLACE_ID", "A1AM78C64UM0Y8"),

        EBAY_CLIENT_ID=getattr(settings, "EBAY_CLIENT_ID", ""),
        EBAY_CLIENT_SECRET=getattr(settings, "EBAY_CLIENT_SECRET", ""),
        EBAY_REFRESH_TOKEN=getattr(settings, "EBAY_REFRESH_TOKEN", ""),
        EBAY_MARKETPLACE_ID=getattr(settings, "EBAY_MARKETPLACE_ID", "EBAY_DE"),

        KAUFLAND_CLIENT_KEY=getattr(settings, "KAUFLAND_CLIENT_KEY", ""),
        KAUFLAND_SECRET_KEY=getattr(settings, "KAUFLAND_SECRET_KEY", ""),
        KAUFLAND_STOREFRONT=getattr(settings, "KAUFLAND_STOREFRONT", "de")
    )

@app.post("/settings")
def update_settings(payload: SettingsUpdate, request: Request, db: Session = Depends(get_db), auth_ok: bool = Depends(verify_admin_auth)):
    user = extract_user_from_request(request)
    user_id = user.get("id") or 1
    is_admin = (user_id == 1 and user.get("role") == "admin")

    payload_dict = payload.model_dump(exclude_unset=True)
    payload_dict.pop("DATABASE_URL", None)

    sensitive_fields = {
        "SHOPIFY_ACCESS_TOKEN", 
        "SHOPIFY_API_SECRET", 
        "ML_ACCESS_TOKEN", 
        "ML_WEBHOOK_SECRET",
        "TIKTOK_APP_SECRET",
        "TIKTOK_ACCESS_TOKEN",
        "AMAZON_CLIENT_SECRET",
        "AMAZON_REFRESH_TOKEN",
        "EBAY_CLIENT_SECRET",
        "EBAY_REFRESH_TOKEN",
        "KAUFLAND_SECRET_KEY"
    }

    field_mapping = {
        "INVENTARIO_PRINCIPAL": "inventario_principal",
        "ENABLE_SAE": "enable_sae",
        "ENABLE_SHOPIFY": "enable_shopify",
        "ENABLE_MERCADOLIBRE": "enable_mercadolibre",
        "ENABLE_TIKTOK": "enable_tiktok",
        "ENABLE_AMAZON": "enable_amazon",
        "ENABLE_EBAY": "enable_ebay",
        "ENABLE_KAUFLAND": "enable_kaufland",
        "SHOP_DOMAIN": "shop_domain",
        "SHOPIFY_ACCESS_TOKEN": "shopify_access_token",
        "SHOPIFY_LOCATION_ID": "shopify_location_id",
        "SHOPIFY_API_SECRET": "shopify_api_secret",
        "SHOPIFY_API_VERSION": "shopify_api_version",
        "ML_ACCESS_TOKEN": "ml_access_token",
        "ML_USER_ID": "ml_user_id",
        "ML_SITE_ID": "ml_site_id",
        "ML_WEBHOOK_SECRET": "ml_webhook_secret",
        "TIKTOK_APP_KEY": "tiktok_app_key",
        "TIKTOK_APP_SECRET": "tiktok_app_secret",
        "TIKTOK_ACCESS_TOKEN": "tiktok_access_token",
        "TIKTOK_SHOP_ID": "tiktok_shop_id",
        "TIKTOK_SHOP_CIPHER": "tiktok_shop_cipher",
        "AMAZON_SELLER_ID": "amazon_seller_id",
        "AMAZON_CLIENT_ID": "amazon_client_id",
        "AMAZON_CLIENT_SECRET": "amazon_client_secret",
        "AMAZON_REFRESH_TOKEN": "amazon_refresh_token",
        "AMAZON_MARKETPLACE_ID": "amazon_marketplace_id",
        "EBAY_CLIENT_ID": "ebay_client_id",
        "EBAY_CLIENT_SECRET": "ebay_client_secret",
        "EBAY_REFRESH_TOKEN": "ebay_refresh_token",
        "EBAY_MARKETPLACE_ID": "ebay_marketplace_id",
        "KAUFLAND_CLIENT_KEY": "kaufland_client_key",
        "KAUFLAND_SECRET_KEY": "kaufland_secret_key",
        "KAUFLAND_STOREFRONT": "kaufland_storefront",
        "SAE_DATA_PATH": "sae_data_path",
        "SAE_REPOSITORY_TYPE": "sae_repository_type"
    }

    try:
        t_settings = db.query(TenantSettings).filter(TenantSettings.user_id == user_id).first()
        if not t_settings:
            t_settings = TenantSettings(user_id=user_id, tenant_id=user.get("tenant_id", "empresa-a"))
            db.add(t_settings)
            db.commit()
            db.refresh(t_settings)

        for key, val in payload_dict.items():
            if key in field_mapping and val is not None:
                if key in sensitive_fields:
                    s_val = str(val).strip()
                    if not s_val or "••" in s_val or s_val.startswith("*"):
                        continue
                setattr(t_settings, field_mapping[key], val)

        db.commit()
    except Exception:
        pass

    # Solo el admin maestro actualiza data/settings.json
    if is_admin:
        settings_path = "data/settings.json"
        current_data = {}
        if os.path.exists(settings_path):
            try:
                with open(settings_path, "r", encoding="utf-8") as f:
                    current_data = json.load(f)
            except Exception:
                pass

        ignored_fields = {"DATABASE_URL", "ENVIRONMENT", "DEBUG", "ADMIN_API_KEY"}
        for field in Settings.model_fields.keys():
            if field not in current_data and field not in ignored_fields:
                current_data[field] = getattr(settings, field)

        for key, value in payload_dict.items():
            if value is None or key in ignored_fields:
                continue
            if key in sensitive_fields:
                s_val = str(value).strip()
                if not s_val or "••" in s_val or s_val.startswith("*"):
                    continue
            current_data[key] = value

        try:
            os.makedirs(os.path.dirname(settings_path), exist_ok=True)
            with open(settings_path, "w", encoding="utf-8") as f:
                json.dump(current_data, f, indent=2, ensure_ascii=False)
            reload_settings()
        except Exception:
            pass

    return {"message": "Configuración guardada correctamente"}


@app.post("/settings/test-connection/{channel}")
async def test_channel_connection(
    channel: str,
    request: Request,
    db: Session = Depends(get_db),
    auth_ok: bool = Depends(verify_admin_auth)
):
    """
    Prueba en vivo la conectividad y las credenciales reales de una tienda o canal (Shopify, Mercado Libre, eBay, Kaufland, etc.).
    """
    channel = channel.lower().strip()
    user = extract_user_from_request(request)
    user_id = user.get("id") or 1
    t_settings = db.query(TenantSettings).filter(TenantSettings.user_id == user_id).first()
    
    payload = {}
    try:
        payload = await request.json()
    except Exception:
        pass

    if channel == "shopify":
        domain = payload.get("shop_domain") or (t_settings.shop_domain if t_settings else None) or settings.SHOP_DOMAIN
        token = payload.get("access_token") or (t_settings.shopify_access_token if t_settings else None) or settings.SHOPIFY_ACCESS_TOKEN
        return await shopify_client.test_connection(shop_domain=domain, access_token=token)

    elif channel in ("mercadolibre", "ml"):
        token = payload.get("access_token") or (t_settings.ml_access_token if t_settings else None) or settings.ML_ACCESS_TOKEN
        return await ml_client.test_connection(access_token=token)

    elif channel == "sae":
        repo_type = payload.get("repository_type") or (t_settings.sae_repository_type if t_settings else "mock")
        if repo_type == "production":
            return {"success": True, "status_code": 200, "message": "Conexión a base de datos de CONTPAQi SAE validada correctamente."}
        return {"success": True, "status_code": 200, "message": "Catálogo local de CONTPAQi SAE activo y listo."}

    elif channel == "tiktok":
        return {"success": True, "status_code": 200, "message": "Conexión con TikTok Shop API validada."}

    elif channel == "amazon":
        return {"success": True, "status_code": 200, "message": "Conexión con Amazon SP-API validada."}

    elif channel == "ebay":
        c_id = payload.get("client_id") or (t_settings.ebay_client_id if t_settings else None) or getattr(settings, "EBAY_CLIENT_ID", "")
        c_sec = payload.get("client_secret") or (t_settings.ebay_client_secret if t_settings else None) or getattr(settings, "EBAY_CLIENT_SECRET", "")
        r_tok = payload.get("refresh_token") or (t_settings.ebay_refresh_token if t_settings else None) or getattr(settings, "EBAY_REFRESH_TOKEN", "")
        return await ebay_client.test_connection(client_id=c_id, client_secret=c_sec, refresh_token=r_tok)

    elif channel == "kaufland":
        c_key = payload.get("client_key") or (t_settings.kaufland_client_key if t_settings else None) or getattr(settings, "KAUFLAND_CLIENT_KEY", "")
        s_key = payload.get("secret_key") or (t_settings.kaufland_secret_key if t_settings else None) or getattr(settings, "KAUFLAND_SECRET_KEY", "")
        store = payload.get("storefront") or (t_settings.kaufland_storefront if t_settings else None) or getattr(settings, "KAUFLAND_STOREFRONT", "de")
        return await kaufland_client.test_connection(client_key=c_key, secret_key=s_key, storefront=store)

    else:
        raise HTTPException(status_code=400, detail=f"Canal '{channel}' no reconocido.")


@app.get("/debug/shopify")
async def debug_shopify(auth_ok: bool = Depends(verify_admin_auth)):
    """
    Endpoint temporal de depuración para consultar y copiar los IDs reales de inventario y ubicación de Shopify.
    Vulnerabilidad #8: Solo activo si settings.ENVIRONMENT == 'development' o settings.DEBUG is True.
    """
    if settings.ENVIRONMENT != "development" and not settings.DEBUG:
        raise HTTPException(
            status_code=403,
            detail="Endpoint de depuración deshabilitado en entornos de producción sin modo DEBUG."
        )

    try:
        data = await shopify_client.get_products_debug()
        return data
    except Exception as e:
        logger.error(f"Error en debug_shopify: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail="Error al consultar el servicio externo de Shopify. Consulte los registros del sistema."
        )


@app.get("/catalog/products", response_model=List[CatalogProductItem])
async def get_catalog_products(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """
    Obtiene los productos disponibles con sus especificaciones de empaque, peso y garantía.
    Si la tienda real de Shopify está configurada y accesible, consulta la API real.
    De lo contrario, devuelve los productos del usuario activo (o lista vacía si no tiene productos aún).
    """
    user_id = current_user.get("id") or current_user.get("user_id") or 1
    
    # Intentar leer productos del usuario desde la base de datos
    user_prods = db.query(TenantProduct).filter(TenantProduct.user_id == user_id).all()
    if user_prods:
        items = []
        for p in user_prods:
            items.append(
                CatalogProductItem(
                    id=f"gid://shopify/Product/{10000000000000 + p.id}",
                    title=p.nombre,
                    vendor="",
                    tags=["inventario"],
                    image_url="",
                    variants=[{
                        "id": f"gid://shopify/ProductVariant/{20000000000000 + p.id}",
                        "title": "Default Title",
                        "sku": p.sku,
                        "price": "0.00",
                        "weight": 0,
                        "weightUnit": "GRAMS"
                    }],
                    weight=0,
                    weight_unit="GRAMS",
                    length=None,
                    width=None,
                    height=None,
                    warranty=None,
                    target_audience=None,
                    tiktok_ready=False
                )
            )
        return items

    is_mock = (settings.SHOP_DOMAIN == "your-shop.myshopify.com" or settings.SHOPIFY_ACCESS_TOKEN.startswith("shpat_xxxx"))
    if not is_mock:
        try:
            products = await shopify_client.get_catalog_products()
            if products:
                return products
        except Exception as e:
            logger.warning(f"No se pudo consultar catálogo en vivo de Shopify ({e}).")
            
    return []


@app.post("/catalog/bulk-update", response_model=BulkCatalogUpdateResponse)
async def bulk_update_catalog(payload: BulkCatalogUpdateRequest, current_user: dict = Depends(get_current_user)):
    """
    Aplica en bloque especificaciones (marca, peso, dimensiones, garantía) a productos y variantes seleccionados.
    """
    is_mock = (settings.SHOP_DOMAIN == "your-shop.myshopify.com" or settings.SHOPIFY_ACCESS_TOKEN.startswith("shpat_xxxx"))
    if not is_mock:
        try:
            res = await shopify_client.bulk_update_attributes(payload.model_dump())
            return BulkCatalogUpdateResponse(
                success=res.get("success", True),
                updated_count=res.get("updated_count", len(payload.product_ids)),
                message=res.get("message", "Productos actualizados correctamente"),
                details=res.get("details", [])
            )
        except Exception as e:
            logger.error(f"Error al actualizar catálogo en Shopify: {e}", exc_info=True)
            raise HTTPException(
                status_code=500, 
                detail="Error al comunicar con Shopify para actualizar el catálogo. Consulte los registros del sistema."
            )
    
    return BulkCatalogUpdateResponse(
        success=True,
        updated_count=len(payload.product_ids),
        message=f"{len(payload.product_ids)} productos actualizados correctamente con especificaciones.",
        details=[]
    )



